# -*- coding: utf-8 -*-
"""
import_utils.py
منطق قراءة ملفات Excel لميزة "الاستيراد والتوزيع التلقائي".

الفكرة:
- نفتح ملف الإكسل المرفوع بواسطة openpyxl.
- نحاول اكتشاف صف العناوين تلقائياً، ثم نحدد أعمدة "اسم الطالب" و"رقم الهوية"
  (وإن وُجدا: "رقم الطالب/الجوال" و"ملاحظات") اعتماداً على نص العنوان، بدلاً
  من افتراض ترتيب أعمدة ثابت.
- إن تعذّر اكتشاف صف عناوين واضح، نرجع تلقائياً لخطة احتياطية بترتيب أعمدة
  ثابت (A=الاسم, B=رقم الهوية, C=رقم الطالب, D=ملاحظات) حتى لا تفشل العملية،
  مع الإفصاح بوضوح للمستخدم عن ذلك في نتيجة التحليل.

هذه الوحدة قراءة فقط: لا تكتب في قاعدة البيانات ولا تُعدّل db.py، وتُستخدم
نتيجتها (rows بصيغة name/national_id/phone/notes) مع الدالة الموجودة أصلاً
db.create_structure_and_distribute_from_rows لتنفيذ التوزيع الفعلي.
"""

import os
import re
import string

import openpyxl

MAX_HEADER_SCAN_ROWS = 15

# كلمات مفتاحية للتعرّف على كل عمود (نصوص مُطبَّعة مسبقاً - انظر normalize_text)
NAME_KEYWORDS = [
    "اسم الطالب", "الاسم الكامل", "اسم كامل", "اسم", "الاسم", "الطالب", "طالب",
    "student name", "full name", "name",
]
ID_KEYWORDS = [
    "رقم الهويه الوطنيه", "رقم الهويه", "الهويه الوطنيه", "الهويه", "هويه",
    "الرقم القومي", "رقم قومي", "السجل المدني", "رقم السجل المدني",
    "رقم الاقامه", "الاقامه",
    "national id", "id number", "identity", "civil id", "national no",
]
PHONE_KEYWORDS = [
    "رقم الطالب", "رقم الجوال", "الجوال", "رقم الهاتف", "الهاتف", "رقم التواصل",
    "الموبايل", "رقم الموبايل",
    "phone", "mobile", "student number", "contact number",
]
NOTES_KEYWORDS = [
    "ملاحظات", "ملاحظه", "ملحوظات", "notes", "note", "remarks",
]

_ARABIC_DIACRITICS = re.compile(r"[\u0610-\u061A\u064B-\u065F\u06D6-\u06ED\u0640]")


def normalize_text(value):
    """تطبيع نص عربي/إنجليزي للمقارنة: إزالة التشكيل والتطويل، توحيد الألف/التاء
    المربوطة، خفض الحالة، وضغط الفراغات. يُستخدم فقط للمطابقة، لا للعرض."""
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = _ARABIC_DIACRITICS.sub("", s)
    s = s.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا").replace("ٱ", "ا")
    s = s.replace("ى", "ي").replace("ة", "ه")
    s = re.sub(r"\s+", " ", s).strip()
    s = s.strip(string.punctuation + "ـ :،-")
    return s


def _cell_to_str(value):
    """يحوّل قيمة خلية Excel إلى نص نظيف، مع تفادي أخطاء الأرقام العشرية
    الشائعة عند قراءة أعمدة أرقام هوية/جوال (مثال: 123456789.0 -> 123456789)."""
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _pick_sheet(wb):
    """يختار أول ورقة عمل تحتوي بيانات فعلية (أكثر من صف واحد أو تحتوي محتوى)."""
    candidates = wb.worksheets
    for ws in candidates:
        if ws.max_row and ws.max_row >= 1 and ws.max_column and ws.max_column >= 1:
            # تأكد من وجود خلية غير فارغة واحدة على الأقل
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
                if any(c.value not in (None, "") for c in row):
                    return ws
    return candidates[0] if candidates else None


def _match_keyword(normalized_cell, keywords):
    for kw in keywords:
        if kw and kw in normalized_cell:
            return True
    return False


def _detect_header_row(ws):
    """يفحص أول عدد من الصفوف بحثاً عن صف عناوين، ويرجع:
    (header_row_index أو None, mapping) حيث mapping = {"name": col, "national_id": col, ...}
    بأرقام أعمدة تبدأ من 1، أو None لأي عمود لم يُكتشف."""
    best_row = None
    best_score = 0
    best_mapping = {}

    max_scan = min(ws.max_row or 1, MAX_HEADER_SCAN_ROWS)
    for r in range(1, max_scan + 1):
        mapping = {}
        score = 0
        for c in range(1, (ws.max_column or 1) + 1):
            cell_val = ws.cell(row=r, column=c).value
            norm = normalize_text(cell_val)
            if not norm:
                continue
            if "name" not in mapping and _match_keyword(norm, NAME_KEYWORDS):
                mapping["name"] = c
                score += 3
                continue
            if "national_id" not in mapping and _match_keyword(norm, ID_KEYWORDS):
                mapping["national_id"] = c
                score += 3
                continue
            if "phone" not in mapping and _match_keyword(norm, PHONE_KEYWORDS):
                mapping["phone"] = c
                score += 2
                continue
            if "notes" not in mapping and _match_keyword(norm, NOTES_KEYWORDS):
                mapping["notes"] = c
                score += 1
                continue
        if score > best_score:
            best_score = score
            best_row = r
            best_mapping = mapping

    if best_score <= 0 or "name" not in best_mapping:
        return None, {}
    return best_row, best_mapping


def _fallback_mapping(ws):
    """خطة احتياطية بترتيب أعمدة ثابت عندما يتعذّر اكتشاف صف عناوين واضح:
    العمود الأول = الاسم، الثاني = رقم الهوية، الثالث = رقم الطالب، الرابع = ملاحظات."""
    ncols = ws.max_column or 1
    mapping = {"name": 1}
    if ncols >= 2:
        mapping["national_id"] = 2
    if ncols >= 3:
        mapping["phone"] = 3
    if ncols >= 4:
        mapping["notes"] = 4
    return mapping


def column_letter(idx):
    return openpyxl.utils.get_column_letter(idx) if idx else ""


def build_column_options(ws, header_row):
    """يبني قائمة الأعمدة المتاحة (لعرضها في واجهة إعادة التعيين اليدوي)،
    كل عنصر: {index, letter, label} حيث label = نص العنوان إن وُجد صف عناوين،
    وإلا معاينة أول قيمة بيانات فعلية في هذا العمود."""
    options = []
    ncols = ws.max_column or 1
    preview_row = header_row + 1 if header_row else 1
    for c in range(1, ncols + 1):
        label = ""
        if header_row:
            label = _cell_to_str(ws.cell(row=header_row, column=c).value)
        if not label:
            label = _cell_to_str(ws.cell(row=preview_row, column=c).value)
        options.append({
            "index": c,
            "letter": column_letter(c),
            "label": label or "(بدون عنوان)",
        })
    return options


def _iter_data_rows(ws, header_row):
    start = (header_row or 0) + 1
    for r in range(start, (ws.max_row or 0) + 1):
        yield r


def _parse_with_mapping(ws, header_row, mapping):
    """يقرأ صفوف البيانات فعلياً حسب mapping المعطى (بأرقام أعمدة)، ويرجع
    (rows, stats) حيث rows = [{name, national_id, phone, notes}, ...]."""
    name_col = mapping.get("name")
    id_col = mapping.get("national_id")
    phone_col = mapping.get("phone")
    notes_col = mapping.get("notes")

    rows = []
    skipped_empty = 0
    skipped_missing_name = 0
    seen_ids = {}
    duplicate_ids = []
    total_scanned = 0

    for r in _iter_data_rows(ws, header_row):
        row_cells = ws[r] if ws.max_column else []
        if not any((c.value not in (None, "")) for c in row_cells):
            continue  # صف فارغ تماماً: نتجاهله بصمت (لا يُحتسب ضمن total_scanned)

        total_scanned += 1
        name_val = _cell_to_str(ws.cell(row=r, column=name_col).value) if name_col else ""
        if not name_val:
            skipped_missing_name += 1
            continue

        national_id = _cell_to_str(ws.cell(row=r, column=id_col).value) if id_col else ""
        phone = _cell_to_str(ws.cell(row=r, column=phone_col).value) if phone_col else ""
        notes = _cell_to_str(ws.cell(row=r, column=notes_col).value) if notes_col else ""

        if national_id:
            if national_id in seen_ids:
                if len(duplicate_ids) < 8:
                    duplicate_ids.append(national_id)
            else:
                seen_ids[national_id] = True

        rows.append({
            "name": name_val,
            "national_id": national_id,
            "phone": phone,
            "notes": notes,
        })

    stats = {
        "total_scanned": total_scanned,
        "skipped_empty": skipped_empty,
        "skipped_missing_name": skipped_missing_name,
        "duplicate_national_ids": duplicate_ids,
        "imported_count": len(rows),
    }
    return rows, stats


def analyze_excel(filepath):
    """يفتح الملف، يكتشف الورقة وصف العناوين والأعمدة تلقائياً، ويقرأ كل
    الصفوف. يرجع dict شامل لعرضه في شاشة المراجعة ولإعادة التحليل لاحقاً
    بتخطيط أعمدة مُعدَّل يدوياً إن احتاج المستخدم لذلك."""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    try:
        ws = _pick_sheet(wb)
        if ws is None:
            raise ValueError("الملف لا يحتوي على أي بيانات قابلة للقراءة")

        header_row, mapping = _detect_header_row(ws)
        detection_mode = "auto"
        if not mapping or "name" not in mapping:
            mapping = _fallback_mapping(ws)
            header_row = None
            detection_mode = "fallback_fixed_positions"

        rows, stats = _parse_with_mapping(ws, header_row, mapping)
        column_options = build_column_options(ws, header_row)
        sheet_name = ws.title
        max_column = ws.max_column or 1
    finally:
        wb.close()

    return {
        "sheet_name": sheet_name,
        "header_row": header_row,
        "detection_mode": detection_mode,
        "mapping": mapping,
        "column_options": column_options,
        "max_column": max_column,
        "rows": rows,
        "stats": stats,
    }


_NATIONAL_ID_DIGITS_RE = re.compile(r"^\d+$")
_MIN_VALID_NAME_LEN = 2


def dedupe_rows(rows):
    """يمنع تكرار الطلاب قبل الحفظ: يُبقي أول ظهور لكل رقم هوية ويستبعد
    التكرارات اللاحقة (الصفوف بلا رقم هوية تُحفظ كلها لأنها غير قابلة
    للمطابقة الموثوقة). يرجع (unique_rows, removed_count)."""
    seen = set()
    unique_rows = []
    removed = 0
    for row in rows:
        national_id = (row.get("national_id") or "").strip()
        if national_id:
            if national_id in seen:
                removed += 1
                continue
            seen.add(national_id)
        unique_rows.append(row)
    return unique_rows, removed


def validate_rows(rows, existing_national_ids=None):
    """يتحقق من صحة كل صف بعد التحليل و**قبل** حفظه فعلياً في قاعدة البيانات
    (تنفيذاً لمتطلب: "التحقق من صحة البيانات المستوردة قبل الحفظ").

    يفحص كل صف من ثلاث زوايا، دون حذف أي صف تلقائياً (فقط إرفاق تحذيرات
    ليقرر المستخدم بنفسه من شاشة المراجعة):
      - اسم قصير جداً (أقل من حرفين) قد يدل على اختيار عمود خاطئ للاسم.
      - رقم هوية غير رقمي بالكامل (يحتوي حروفاً/رموزاً غير متوقعة).
      - رقم هوية يطابق طالباً نشطاً موجوداً فعلاً في قاعدة البيانات الحالية
        (تكرار حقيقي عبر النظام، وليس فقط داخل الملف نفسه).

    rows: نتيجة _parse_with_mapping (قائمة قواميس name/national_id/phone/notes).
    existing_national_ids: مجموعة أرقام هوية الطلاب النشطين الحاليين في قاعدة
        البيانات (يمررها app.py عبر db.py دون أن تحتاج هذه الوحدة الاتصال
        بقاعدة البيانات مباشرة، حفاظاً على فصل المسؤوليات).
    يرجع (rows_with_warnings, validation_stats)."""
    existing_national_ids = existing_national_ids or set()
    invalid_id_count = 0
    short_name_count = 0
    existing_dup_count = 0
    existing_dup_samples = []

    annotated = []
    for row in rows:
        warnings = []
        name = row.get("name", "") or ""
        national_id = row.get("national_id", "") or ""

        if len(name.strip()) < _MIN_VALID_NAME_LEN:
            warnings.append("short_name")
            short_name_count += 1

        if national_id and not _NATIONAL_ID_DIGITS_RE.match(national_id):
            warnings.append("invalid_national_id_format")
            invalid_id_count += 1

        if national_id and national_id in existing_national_ids:
            warnings.append("exists_in_db")
            existing_dup_count += 1
            if len(existing_dup_samples) < 8:
                existing_dup_samples.append(national_id)

        annotated.append({**row, "warnings": warnings})

    validation_stats = {
        "invalid_national_id_count": invalid_id_count,
        "short_name_count": short_name_count,
        "existing_in_db_count": existing_dup_count,
        "existing_in_db_samples": existing_dup_samples,
    }
    return annotated, validation_stats


def parse_with_manual_mapping(filepath, header_row, name_col, national_id_col=None,
                               phone_col=None, notes_col=None):
    """يعيد القراءة من نفس الملف باستخدام تخطيط أعمدة اختاره المستخدم يدوياً
    (مثلاً بعد أن لاحظ أن الاكتشاف التلقائي اختار عموداً غير صحيح)."""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    try:
        ws = _pick_sheet(wb)
        mapping = {"name": int(name_col)}
        if national_id_col:
            mapping["national_id"] = int(national_id_col)
        if phone_col:
            mapping["phone"] = int(phone_col)
        if notes_col:
            mapping["notes"] = int(notes_col)
        rows, stats = _parse_with_mapping(ws, header_row, mapping)
        column_options = build_column_options(ws, header_row)
        sheet_name = ws.title
    finally:
        wb.close()

    return {
        "sheet_name": sheet_name,
        "header_row": header_row,
        "detection_mode": "manual",
        "mapping": mapping,
        "column_options": column_options,
        "rows": rows,
        "stats": stats,
    }
