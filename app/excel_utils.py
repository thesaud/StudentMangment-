# -*- coding: utf-8 -*-
"""
excel_utils.py — تصدير بيانات الطلاب/الطوابير/المخالفات إلى Excel.

ترتيب الأعمدة الموحّد (يطابق الواجهة في كل الجداول والتصديرات):
    #  |  اسم الطالب  |  رقم الهوية  |  رقم الطالب  |  الكتيبة  |  السرية  |  الفصيل  |  ملاحظات

هذا الملف هو المكان الوحيد لتعريف ترتيب الأعمدة — كل الدوال تستخدمه.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

from . import db

# ---------------------------------------------------------------------------
# أنماط مشتركة
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12, name="Cairo")
TITLE_FONT = Font(bold=True, size=16, color="1F4E78", name="Cairo")
SUBTITLE_FONT = Font(bold=True, size=11, color="555555", name="Cairo")
SECTION_FONT = Font(bold=True, size=13, color="1F4E78", name="Cairo")
LABEL_FONT = Font(bold=True, size=11, name="Cairo")
CELL_FONT = Font(size=11, name="Cairo")
THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

# ---------------------------------------------------------------------------
# أعمدة موحّدة — تعريف واحد يُستخدم في كل مكان
# ---------------------------------------------------------------------------

# أعمدة جدول الطلاب الكامل (يطابق ترتيب الواجهة بالضبط)
STUDENT_FULL_HEADERS = ["#", "اسم الطالب", "رقم الهوية", "رقم الطالب",
                        "الكتيبة", "السرية", "الفصيل", "ملاحظات"]
STUDENT_FULL_WIDTHS = [6, 24, 16, 12, 16, 14, 14, 26]


def _safe(val):
    """يحوّل None إلى سلسلة فارغة."""
    return "" if val is None else val


def _get(obj, key, default=""):
    """قراءة آمنة من dict أو sqlite3.Row."""
    try:
        keys = obj.keys()
        return obj[key] if key in keys else default
    except (AttributeError, TypeError):
        return obj.get(key, default) if obj else default


def _student_number(s):
    """يقرأ رقم الطالب التسلسلي من student_number أولاً، ثم phone كاحتياطي."""
    sn = _get(s, "student_number")
    if sn:
        return str(sn)
    return _safe(_get(s, "phone"))


def _student_row(s, idx, battalion="", company="", platoon=""):
    """صف طالب بالترتيب الموحّد: #, Name, NatID, StudentNum, Bat, Co, Pl, Notes."""
    return [
        idx,
        _safe(_get(s, "name") or _get(s, "sname")),
        _safe(_get(s, "national_id")),
        _student_number(s),
        _safe(battalion or _get(s, "bname")),
        _safe(company or _get(s, "cname")),
        _safe(platoon or _get(s, "pname")),
        _safe(_get(s, "notes")),
    ]


# ---------------------------------------------------------------------------
# أدوات تنسيق مشتركة
# ---------------------------------------------------------------------------

def _style_header_row(ws, row_idx, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _style_data_row(ws, row_idx, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.border = BORDER
        cell.alignment = RIGHT
        cell.font = CELL_FONT


def _auto_width(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def _write_title(ws, text, ncols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.cell(row=row, column=1, value=text).font = TITLE_FONT
    ws.cell(row=row, column=1).alignment = CENTER


def _write_subtitle(ws, text, ncols, row=2):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.cell(row=row, column=1, value=text).font = SUBTITLE_FONT
    ws.cell(row=row, column=1).alignment = CENTER


# ---------------------------------------------------------------------------
# 1) تصدير الهيكل الكامل
# ---------------------------------------------------------------------------

def export_full_hierarchy(filepath, tree):
    """كامل الهيكل التنظيمي بالترتيب الموحّد."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "جميع الطلاب"
    ws.sheet_view.rightToLeft = True

    h = STUDENT_FULL_HEADERS
    ws.append(h)
    _style_header_row(ws, 1, len(h))

    idx = 0
    for b in tree:
        for c in b["companies"]:
            for p in c["platoons"]:
                for s in p.get("students", []):
                    idx += 1
                    ws.append(_student_row(s, idx, b["row"]["name"], c["row"]["name"], p["row"]["name"]))
                    _style_data_row(ws, ws.max_row, len(h))

    _auto_width(ws, STUDENT_FULL_WIDTHS)
    wb.save(filepath)


# ---------------------------------------------------------------------------
# 2) تصدير فصيل
# ---------------------------------------------------------------------------

def export_platoon_students(filepath, platoon_name, company_name, battalion_name, students):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الطلاب"
    ws.sheet_view.rightToLeft = True

    h = STUDENT_FULL_HEADERS
    _write_title(ws, f"كشف طلاب فصيل: {platoon_name}  |  سرية: {company_name}  |  كتيبة: {battalion_name}", len(h))
    ws.append([])
    ws.append(h)
    _style_header_row(ws, 3, len(h))

    for i, s in enumerate(students, start=1):
        ws.append(_student_row(s, i, battalion_name, company_name, platoon_name))
        _style_data_row(ws, ws.max_row, len(h))

    _auto_width(ws, STUDENT_FULL_WIDTHS)
    wb.save(filepath)


# ---------------------------------------------------------------------------
# 3) تصدير سرية
# ---------------------------------------------------------------------------

def export_company_students(filepath, company_name, battalion_name, platoons):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "طلاب السرية"
    ws.sheet_view.rightToLeft = True

    h = STUDENT_FULL_HEADERS
    _write_title(ws, f"كشف طلاب سرية: {company_name}  |  كتيبة: {battalion_name}", len(h))
    ws.append([])
    ws.append(h)
    _style_header_row(ws, 3, len(h))

    idx = 0
    for p in platoons:
        for s in p["students"]:
            idx += 1
            ws.append(_student_row(s, idx, battalion_name, company_name, p["name"]))
            _style_data_row(ws, ws.max_row, len(h))

    _auto_width(ws, STUDENT_FULL_WIDTHS)
    wb.save(filepath)


# ---------------------------------------------------------------------------
# 4) تصدير كتيبة
# ---------------------------------------------------------------------------

def export_battalion_students(filepath, battalion_name, companies):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "طلاب الكتيبة"
    ws.sheet_view.rightToLeft = True

    h = STUDENT_FULL_HEADERS
    _write_title(ws, f"كشف طلاب كتيبة: {battalion_name}", len(h))
    ws.append([])
    ws.append(h)
    _style_header_row(ws, 3, len(h))

    idx = 0
    for c in companies:
        for p in c["platoons"]:
            for s in p["students"]:
                idx += 1
                ws.append(_student_row(s, idx, battalion_name, c["name"], p["name"]))
                _style_data_row(ws, ws.max_row, len(h))

    _auto_width(ws, STUDENT_FULL_WIDTHS)
    wb.save(filepath)


# ---------------------------------------------------------------------------
# 5) تصدير المستقيلين
# ---------------------------------------------------------------------------

def export_resigned_students(filepath, resigned_students):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "المستقيلون"
    ws.sheet_view.rightToLeft = True

    headers = ["#", "اسم الطالب", "رقم الهوية", "رقم الطالب",
               "الكتيبة", "السرية", "الفصيل",
               "سبب الاستقالة", "تاريخ الاستقالة"]
    widths = [6, 22, 16, 12, 16, 14, 14, 26, 16]

    _write_title(ws, "جدول الطلاب المستقيلين", len(headers))
    _write_subtitle(ws, f"تاريخ الطباعة: {datetime.now().strftime('%Y-%m-%d %H:%M')}", len(headers))
    ws.append([])
    ws.append(headers)
    _style_header_row(ws, 4, len(headers))

    for i, s in enumerate(resigned_students, start=1):
        ws.append([
            i,
            _safe(_get(s, "sname")),
            _safe(_get(s, "national_id")),
            _student_number(s),
            _safe(_get(s, "bname")),
            _safe(_get(s, "cname")),
            _safe(_get(s, "pname")),
            _safe(_get(s, "resignation_reason")),
            _safe(_get(s, "resignation_date")),
        ])
        _style_data_row(ws, ws.max_row, len(headers))

    _auto_width(ws, widths)
    wb.save(filepath)


# ---------------------------------------------------------------------------
# 6) تصدير طابور
# ---------------------------------------------------------------------------

def export_queue(filepath, queue_name, queue_date, students):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الطابور"
    ws.sheet_view.rightToLeft = True

    headers = ["#", "اسم الطالب", "رقم الهوية", "رقم الطالب",
               "الكتيبة", "السرية", "الفصيل",
               "نوع المخالفة", "مدة المخالفة", "الوقت المتبقي", "تاريخ الانتهاء"]
    widths = [6, 24, 16, 12, 16, 14, 14, 18, 16, 16, 16]

    _write_title(ws, f"كشف طابور إضافي: {queue_name}", len(headers))
    _write_subtitle(ws, f"التاريخ: {queue_date}   |   تاريخ الطباعة: {datetime.now().strftime('%Y-%m-%d %H:%M')}", len(headers))
    ws.append([])
    ws.append(headers)
    _style_header_row(ws, 4, len(headers))

    for i, s in enumerate(students, start=1):
        duration_label = db.format_duration_label(_get(s, "q_duration_category"), _get(s, "q_duration_days"))
        remaining_label = db.compute_remaining_label(_get(s, "q_expires_at"), _get(s, "q_duration_category"))
        ws.append([
            i,
            _safe(_get(s, "sname")),
            _safe(_get(s, "national_id")),
            _student_number(s),
            _safe(_get(s, "bname")),
            _safe(_get(s, "cname")),
            _safe(_get(s, "pname")),
            _safe(_get(s, "vtype") or _get(s, "q_violation_type")),
            duration_label,
            remaining_label,
            _safe(_get(s, "q_expires_at")),
        ])
        _style_data_row(ws, ws.max_row, len(headers))

    _auto_width(ws, widths)
    wb.save(filepath)


# ---------------------------------------------------------------------------
# 7) تصدير سجل المخالفات
# ---------------------------------------------------------------------------

def export_violation_catalog(filepath, catalog):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "سجل المخالفات"
    ws.sheet_view.rightToLeft = True

    headers = ["#", "اسم الطالب", "رقم الهوية", "رقم الطالب",
               "الكتيبة", "السرية", "الفصيل",
               "نوع المخالفة", "المدة", "التاريخ", "ملاحظات"]
    widths = [6, 22, 16, 12, 16, 14, 14, 20, 16, 14, 24]

    _write_title(ws, "سجل المخالفات", len(headers))
    _write_subtitle(ws, f"تاريخ الطباعة: {datetime.now().strftime('%Y-%m-%d %H:%M')}", len(headers))
    ws.append([])
    ws.append(headers)
    _style_header_row(ws, 4, len(headers))

    for i, v in enumerate(catalog, start=1):
        ws.append([
            i,
            _safe(_get(v, "sname")),
            _safe(_get(v, "national_id")),
            _student_number(v),
            _safe(_get(v, "bname")),
            _safe(_get(v, "cname")),
            _safe(_get(v, "pname")),
            _safe(_get(v, "violation_type")),
            _safe(_get(v, "duration")),
            _safe(_get(v, "date_added")),
            _safe(_get(v, "notes")),
        ])
        _style_data_row(ws, ws.max_row, len(headers))

    _auto_width(ws, widths)
    wb.save(filepath)


# ---------------------------------------------------------------------------
# 8) تصدير استعلام طالب
# ---------------------------------------------------------------------------

def export_student_inquiry(filepath, student_info, violations, queues):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "استعلام عن طالب"
    ws.sheet_view.rightToLeft = True

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    _write_title(ws, f"تقرير استعلام عن الطالب: {_get(student_info, 'sname')}", 6)
    _write_subtitle(ws, f"تاريخ ووقت الطباعة: {now_str}", 6)

    row = 4
    info_fields = [
        ("اسم الطالب", _get(student_info, "sname")),
        ("رقم الهوية", _get(student_info, "national_id")),
        ("رقم الطالب", _student_number(student_info)),
        ("الكتيبة", _get(student_info, "bname")),
        ("السرية", _get(student_info, "cname")),
        ("الفصيل", _get(student_info, "pname")),
        ("ملاحظات", _get(student_info, "notes")),
    ]
    for label, value in info_fields:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).alignment = RIGHT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.cell(row=row, column=2, value=_safe(value)).font = CELL_FONT
        ws.cell(row=row, column=2).alignment = RIGHT
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    # المخالفات السابقة
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="المخالفات السابقة").font = SECTION_FONT
    ws.cell(row=row, column=1).alignment = RIGHT
    row += 1

    v_headers = ["#", "نوع المخالفة", "المدة", "تاريخ التسجيل", "تاريخ الانتهاء", "ملاحظات"]
    for col, h in enumerate(v_headers, start=1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(v_headers))
    row += 1

    if violations:
        for i, v in enumerate(violations, start=1):
            vals = [i, _get(v, "violation_type"), _get(v, "duration"),
                    _get(v, "date_added"), _get(v, "expires_at"), _get(v, "notes")]
            for col, val in enumerate(vals, start=1):
                c = ws.cell(row=row, column=col, value=_safe(val))
                c.border = BORDER
                c.alignment = RIGHT
                c.font = CELL_FONT
            row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="لا توجد مخالفات مسجّلة لهذا الطالب.").alignment = RIGHT
        row += 1

    # سجل الطوابير الإضافية
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value="سجل الطوابير الإضافية").font = SECTION_FONT
    ws.cell(row=row, column=1).alignment = RIGHT
    row += 1

    q_headers = ["#", "اسم الطابور", "تاريخ الطابور", "مدة المخالفة", "الحالة", "ملاحظات"]
    for col, h in enumerate(q_headers, start=1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(q_headers))
    row += 1

    if queues:
        for i, q in enumerate(queues, start=1):
            duration_label = db.format_duration_label(_get(q, "q_duration_category"), _get(q, "q_duration_days"))
            status_label = "نشط حالياً" if _get(q, "q_is_active", 1) else "منتهٍ / أُزيل"
            vals = [i, _get(q, "qname"), _get(q, "queue_date"), duration_label, status_label, _get(q, "qnotes")]
            for col, val in enumerate(vals, start=1):
                c = ws.cell(row=row, column=col, value=_safe(val))
                c.border = BORDER
                c.alignment = RIGHT
                c.font = CELL_FONT
            row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value="لم يُدرج هذا الطالب في أي طابور إضافي.").alignment = RIGHT
        row += 1

    _auto_width(ws, [6, 22, 16, 16, 16, 26])
    wb.save(filepath)
